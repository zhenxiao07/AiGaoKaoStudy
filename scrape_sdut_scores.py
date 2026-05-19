"""
山东理工大学近五年各专业录取分数爬虫
数据来源：
  - 2020-2022: 山东理工大学招生信息网官方公布数据
  - 2023-2025: 大学生必备网（dxsbb.com）汇总数据
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

requests.packages.urllib3.disable_warnings()

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SESSION = requests.Session()
SESSION.verify = False
SESSION.headers.update(HEADERS)


# ─────────────────────────────────────────────────────────
# 官网数据抓取（2020-2022）
# ─────────────────────────────────────────────────────────

OFFICIAL_URLS = {
    2022: "https://zszx.sdut.edu.cn/f/newsCenter/article/0cab4285593c40ccb75c4cdd44780f64",
    2021: "https://zszx.sdut.edu.cn/f/newsCenter/article/46b1cde977e84c61a14d098cebd06fd8",
    2020: "https://zszx.sdut.edu.cn/f/newsCenter/article/7eda321e0c024b2ab5478cf952d288c0",
}


def fetch_official(year: int, url: str) -> pd.DataFrame:
    print(f"  抓取官网 {year} 年数据 ...")
    try:
        r = SESSION.get(url, timeout=20)
        r.encoding = "utf-8"
        soup = BeautifulSoup(r.text, "html.parser")
        tables = soup.find_all("table")
        if not tables:
            print(f"    ⚠ {year} 年页面无表格")
            return pd.DataFrame()

        rows = []
        for table in tables:
            for tr in table.find_all("tr"):
                cols = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if cols:
                    rows.append(cols)

        if not rows:
            return pd.DataFrame()

        # 取首行作为列名
        raw_header = rows[0]
        # 归一化列名
        col_map = {
            "省份": "省份", "ʡ��": "省份",
            "科类": "科类", "����": "科类",
            "批次": "批次", "����": "批次",
            "录取专业": "专业", "¼ȡרҵ": "专业",
            "人数": "录取人数", "����": "录取人数",
            "最高分": "最高分", "��߷�": "最高分",
            "最低分": "最低分", "��ͷ�": "最低分",
            "平均分": "平均分", "ƽ����": "平均分",
        }
        header = [col_map.get(c, c) for c in raw_header]
        data_rows = rows[1:]

        df = pd.DataFrame(data_rows, columns=header[:len(data_rows[0])] if data_rows else header)
        df.insert(0, "年份", str(year))

        # 过滤掉重复表头行
        df = df[df["专业"] != "录取专业"]
        df = df[df["专业"].notna() & (df["专业"] != "")]

        # 统一列顺序
        want_cols = ["年份", "省份", "科类", "批次", "专业", "最低分", "平均分", "最高分", "录取人数"]
        for c in want_cols:
            if c not in df.columns:
                df[c] = ""
        df = df[want_cols]
        df = df.reset_index(drop=True)
        print(f"    ✓ {year} 年：{len(df)} 条记录")
        return df

    except Exception as e:
        print(f"    ✗ {year} 年抓取失败: {e}")
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────
# 第三方数据抓取（2023-2025）
# ─────────────────────────────────────────────────────────

DXSBB_URL = "https://www.dxsbb.com/news/33038.html"
# 页面中表格顺序：2025, 2024, 2023
DXSBB_TABLE_YEARS = [2025, 2024, 2023]


def fetch_dxsbb() -> pd.DataFrame:
    print("  抓取 dxsbb.com 2023-2025 年数据 ...")
    try:
        r = SESSION.get(DXSBB_URL, timeout=25)
        r.encoding = "utf-8"
        soup = BeautifulSoup(r.text, "html.parser")
        tables = soup.find_all("table")

        all_dfs = []
        for idx, (table, year) in enumerate(zip(tables, DXSBB_TABLE_YEARS)):
            rows = []
            for tr in table.find_all("tr"):
                cols = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if cols:
                    rows.append(cols)
            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:]
            if not data_rows:
                continue

            df = pd.DataFrame(data_rows, columns=header[:len(data_rows[0])])

            # 列名归一化（来源列名：年份,省市,科类,类型,专业,最低分,平均分,最高分,控制线,最低分排名）
            rename_map = {
                "省市": "省份", "类型": "批次",
                "最低分排名": "最低位次", "控制线": "控制线",
            }
            df = df.rename(columns=rename_map)

            # 强制年份字段（有些行自带年份列）
            if "年份" not in df.columns:
                df.insert(0, "年份", str(year))
            else:
                df["年份"] = str(year)

            want_cols = ["年份", "省份", "科类", "批次", "专业", "最低分", "平均分", "最高分", "控制线", "最低位次"]
            for c in want_cols:
                if c not in df.columns:
                    df[c] = ""
            df = df[want_cols]

            # 过滤空行和无意义行
            df = df[df["专业"].notna() & (df["专业"].str.strip() != "")]
            df = df[df["专业"] != "专业"]
            df = df.reset_index(drop=True)

            all_dfs.append(df)
            print(f"    ✓ {year} 年：{len(df)} 条记录")

        if all_dfs:
            return pd.concat(all_dfs, ignore_index=True)
        return pd.DataFrame()

    except Exception as e:
        print(f"    ✗ dxsbb 抓取失败: {e}")
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────
# 合并数据
# ─────────────────────────────────────────────────────────

def merge_data(official_dfs: list, dxsbb_df: pd.DataFrame) -> pd.DataFrame:
    """合并官网数据和第三方数据，统一列结构"""
    all_frames = []

    # 官网数据已有标准列
    for df in official_dfs:
        if df.empty:
            continue
        # 补充缺少的列
        for col in ["控制线", "最低位次"]:
            if col not in df.columns:
                df[col] = ""
        all_frames.append(df)

    # 第三方数据
    if not dxsbb_df.empty:
        # 补充录取人数列
        if "录取人数" not in dxsbb_df.columns:
            dxsbb_df["录取人数"] = ""
        all_frames.append(dxsbb_df)

    if not all_frames:
        return pd.DataFrame()

    # 统一最终列
    final_cols = ["年份", "省份", "科类", "批次", "专业", "最低分", "平均分", "最高分", "录取人数", "控制线", "最低位次"]
    merged = pd.concat(all_frames, ignore_index=True)
    for c in final_cols:
        if c not in merged.columns:
            merged[c] = ""
    merged = merged[final_cols]
    merged = merged.sort_values(["年份", "省份", "专业"]).reset_index(drop=True)
    return merged


# ─────────────────────────────────────────────────────────
# Excel 导出
# ─────────────────────────────────────────────────────────

def style_ws(ws, header_row_idx: int = 1):
    """给工作表应用样式"""
    # 颜色定义
    header_fill = PatternFill("solid", fgColor="2B5FA8")  # 深蓝标题
    odd_fill = PatternFill("solid", fgColor="EEF4FF")    # 浅蓝奇数行
    even_fill = PatternFill("solid", fgColor="FFFFFF")   # 白色偶数行
    year_fills = {
        "2020": PatternFill("solid", fgColor="FFF3CD"),
        "2021": PatternFill("solid", fgColor="D4EDDA"),
        "2022": PatternFill("solid", fgColor="D1ECF1"),
        "2023": PatternFill("solid", fgColor="F8D7DA"),
        "2024": PatternFill("solid", fgColor="E8D5F7"),
        "2025": PatternFill("solid", fgColor="FFDAB9"),
    }
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="微软雅黑", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), start=1):
        for cell in row:
            cell.border = border
            if row_idx == header_row_idx:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            else:
                year_val = ws.cell(row=row_idx, column=1).value
                fill = year_fills.get(str(year_val), (odd_fill if row_idx % 2 else even_fill))
                cell.fill = fill
                cell.font = cell_font
                if cell.column == 5:  # 专业列左对齐
                    cell.alignment = left
                else:
                    cell.alignment = center

    # 冻结首行
    ws.freeze_panes = ws["A2"]
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 列宽
    col_widths = [8, 12, 12, 18, 38, 8, 8, 8, 10, 8, 12]
    for i, w in enumerate(col_widths[:max_col], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22


def export_excel(merged: pd.DataFrame, out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet1: 全部数据汇总 ──
    ws_all = wb.create_sheet("全部数据汇总")
    ws_all.append(list(merged.columns))
    for _, row in merged.iterrows():
        ws_all.append(list(row))
    style_ws(ws_all)
    print(f"  ✓ 全部数据汇总: {len(merged)} 行")

    # ── Sheet2-N: 按年份分Sheet ──
    for year in sorted(merged["年份"].unique()):
        df_year = merged[merged["年份"] == year]
        ws = wb.create_sheet(f"{year}年录取数据")
        ws.append(list(df_year.columns))
        for _, row in df_year.iterrows():
            ws.append(list(row))
        style_ws(ws)
        print(f"  ✓ {year}年: {len(df_year)} 行")

    # ── 山东省专业分数趋势表 ──
    shandong = merged[merged["省份"].isin(["山东", "山东省"])].copy()
    if not shandong.empty:
        ws_sd = wb.create_sheet("山东省专业分数趋势")
        # 透视：专业 × 年份 的最低分
        pivot = shandong.pivot_table(
            index="专业", columns="年份", values="最低分", aggfunc="min"
        )
        pivot = pivot.reset_index()
        pivot.columns = [str(c) for c in pivot.columns]
        ws_sd.append(list(pivot.columns))
        for _, row in pivot.iterrows():
            ws_sd.append(list(row))
        style_ws(ws_sd)
        print(f"  ✓ 山东省专业分数趋势: {len(pivot)} 个专业")

    wb.save(out_path)
    print(f"\n✅ Excel 已保存至: {out_path}")


# ─────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  山东理工大学近五年各专业录取分数爬取")
    print("=" * 60)

    # 1. 抓取官网 2020-2022
    print("\n[1/2] 抓取官方招生网数据（2020-2022）")
    official_dfs = []
    for year, url in OFFICIAL_URLS.items():
        df = fetch_official(year, url)
        official_dfs.append(df)
        time.sleep(1.5)

    # 2. 抓取第三方 2023-2025
    print("\n[2/2] 抓取第三方汇总数据（2023-2025）")
    dxsbb_df = fetch_dxsbb()

    # 3. 合并
    print("\n[合并] 整合所有数据...")
    merged = merge_data(official_dfs, dxsbb_df)
    print(f"  总计：{len(merged)} 条录取记录，覆盖 {sorted(merged['年份'].unique())} 年")

    # 4. 导出
    out_path = r"C:\Users\l\Desktop\AI智能填报志愿\山东理工大学近五年录取分数线.xlsx"
    print(f"\n[导出] 写入 Excel ...")
    export_excel(merged, out_path)

    print("\n数据结构预览（前5行）：")
    print(merged.head().to_string(index=False))


if __name__ == "__main__":
    main()
